import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Define user mapping
valid_users = {
    "abharti": "Ankur", "agiri1": "Aman", "dahuja": "Daksh", "dmam": "Deepak",
    "mranganathan": "Magesh", "psrihari": "Prakasam", "rjain6": "Rohit",
    "sarikapudi": "Sudheer", "sjain16": "Siddharth", "spatnam": "Sreekanth"
}

def process_excel(file):
    df = pd.read_excel(file, engine='pyxlsb')
    
    # Apply filters
    filtered_df = df[
        (df["QUEUE_CODE"] == "BDWCNFG") &
        (df["D_IN_OUT_OLA"] == "OUT OF OLA") &
        (df["USER_ID_COMPLETION"].isin(valid_users.keys()))
    ].copy()
    
    # Add Failure category & Failure Reasons
    if "Failure category" not in filtered_df.columns:
        filtered_df["Failure category"] = ""
    if "Failure Reasons" not in filtered_df.columns:
        filtered_df["Failure Reasons"] = ""
    
    filtered_df.loc[filtered_df["DELAY_DIARY"].isna(), "Failure category"] = "Genuine Fault / Prioritization Error"
    filtered_df.loc[filtered_df["DELAY_DIARY"].isna(), "Failure Reasons"] = filtered_df["USER_ID_COMPLETION"].map(
        lambda x: f"Missed to close on time by {valid_users.get(x, x)}"
    )
    
    return filtered_df

def format_excel(df):
    wb = Workbook()
    ws = wb.active

    selected_columns = [
        "QUEUE_CODE", "TASK_CLOSED", "NEW_CONTRACT_NO", "COUNTRY", "WORK_ITEM_ID_CALC", "REPORTING_WEEK", 
        "PRODUCT_OFFERING", "D_OLA_TARGET", "LEAD_TIME_OVERALL", "D_IN_OUT_OLA", "USER_ID_COMPLETION", 
        "CUSTOMER_NAME", "Sub Team", "Failure category", "Failure Reasons"
    ]
    df = df[selected_columns]
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    font = Font(name='Arial', size=9)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal='left', vertical='center')
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = font
            cell.border = border
            cell.alignment = alignment
            if cell.row == 1:
                cell.fill = header_fill
    
    return wb

def generate_email_text(df, weeks):
    email_text = f"Hi Team,\n\nBelow is the OLA Analysis report for Week {', '.join(map(str, weeks))}.\n\n"
    email_text += df.to_csv(sep='\t', index=False)
    return email_text

# Streamlit UI
st.title("OLA Data Processor")
uploaded_files = st.file_uploader("Upload Excel Files", type=["xlsb"], accept_multiple_files=True)

if uploaded_files:
    st.write("Processing files...")
    
    filtered_dfs = []
    reporting_weeks = []
    
    for uploaded_file in uploaded_files:
        filtered_df = process_excel(uploaded_file)
        filtered_dfs.append(filtered_df)
        reporting_weeks.extend(filtered_df["REPORTING_WEEK"].unique().tolist())
    
    if filtered_dfs:
        consolidated_df = pd.concat(filtered_dfs, ignore_index=True)
        
        st.write("### Consolidated Filtered Data Preview:")
        st.dataframe(consolidated_df)
        
        wb = format_excel(consolidated_df)
        output_filename = "consolidated_filtered_data.xlsx"
        wb.save(output_filename)

        with open(output_filename, "rb") as file:
            st.download_button(
                label="Download Consolidated Excel File",
                data=file,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        email_text = generate_email_text(consolidated_df, sorted(set(reporting_weeks)))
        st.write("### Email Text Preview:")
        st.text_area("Copy and paste this email:", email_text, height=300)
